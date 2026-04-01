"""
advanced_excel.py
-----------------
Python tool for AI agents to interact with the LBO Excel model (.xlsx).

Usage:
    from advanced_excel import AdvancedExcel

    tool = AdvancedExcel("IOI Model Template.xlsx", "output/LBO_populated.xlsx")
    tool.write_cell("Model", "H5", "AcmeCorp")
    tool.write_cell("Model", "I109", 85.0)
    tool.format_cell("Model", "I109", number_format='#,##0.0"x"')
    tool.save()
    print(tool.get_audit_log())

Sheets in IOI Model Template.xlsx:
    - Model              (main input sheet, A2:AU398)
    - Output AVP         (output / print sheet, B2:AF87)
    - P&L (presentation) (presentation income statement, A2:S37)
    - PB_CACHE           (internal cache, do not touch)
"""

import copy
import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Optional, Union

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string

# Sheets that agents should never touch
_PROTECTED_SHEETS = {"PB_CACHE"}

# Valid sheet names in this workbook
_VALID_SHEETS = {"Output AVP", "PB_CACHE", "Model", "P&L (presentation)"}


class ExcelToolError(Exception):
    """Raised for validation errors, formula protection violations, etc."""


class AdvancedExcel:
    """
    Tool for AI agents to read, write, and format cells in the LBO Excel model.

    Parameters
    ----------
    template_path : str | Path
        Path to the source .xlsx template (never modified).
    output_path : str | Path
        Path where the working copy will be saved. Created on first save().
    agent_id : str, optional
        Identifier for the calling agent, recorded in the audit log.
    """

    def __init__(
        self,
        template_path: Union[str, Path] = "IOI Model Template.xlsx",
        output_path: Union[str, Path] = "output/LBO_populated.xlsx",
        agent_id: str = "agent",
    ):
        self.template_path = Path(template_path)
        self.output_path = Path(output_path)
        self.agent_id = agent_id
        self._audit: list[dict] = []

        if not self.template_path.exists():
            raise FileNotFoundError(f"Template not found: {self.template_path}")

        # Load a working copy into memory (keeps template untouched on disk)
        self._wb = load_workbook(self.template_path, data_only=False)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def read_cell(self, sheet: str, cell: str) -> Any:
        """
        Read the value of a single cell.

        Parameters
        ----------
        sheet : str
            Sheet name (e.g. "Model").
        cell : str
            Cell address (e.g. "H5").

        Returns
        -------
        The cell's value (could be a number, string, datetime, or formula string).
        """
        ws = self._get_sheet(sheet)
        self._validate_cell_address(cell)
        value = ws[cell].value
        self._log("read", sheet, cell, value, None)
        return value

    def read_range(self, sheet: str, cell_range: str) -> list[list[Any]]:
        """
        Read a rectangular range of cells.

        Parameters
        ----------
        sheet : str
            Sheet name.
        cell_range : str
            Range address (e.g. "B5:H20").

        Returns
        -------
        List of rows, each row is a list of cell values.
        """
        ws = self._get_sheet(sheet)
        result = []
        for row in ws[cell_range]:
            result.append([cell.value for cell in row])
        self._log("read_range", sheet, cell_range, result, None)
        return result

    def write_cell(
        self,
        sheet: str,
        cell: str,
        value: Any,
        force: bool = False,
    ) -> None:
        """
        Write a value to a cell. Raises ExcelToolError if the cell currently
        holds a formula (use write_formula to override intentionally, or
        pass force=True to bypass the guard).

        Parameters
        ----------
        sheet : str
            Sheet name.
        cell : str
            Cell address (e.g. "H5").
        value : Any
            The value to write (str, int, float, datetime, None).
        force : bool
            If True, overwrite formula cells without raising an error.
            Use sparingly — prefer write_formula for intentional formula changes.
        """
        ws = self._get_sheet(sheet)
        self._validate_cell_address(cell)
        self._check_not_protected_sheet(sheet)

        current = ws[cell].value
        if isinstance(current, str) and current.startswith("=") and not force:
            raise ExcelToolError(
                f"Cell {sheet}!{cell} contains a formula ({current!r}). "
                "Use write_formula() to write a formula, or pass force=True to "
                "overwrite with a hard-coded value."
            )

        ws[cell] = value
        self._log("write", sheet, cell, current, value)

    def write_formula(self, sheet: str, cell: str, formula: str) -> None:
        """
        Write an Excel formula into a cell.

        Parameters
        ----------
        sheet : str
            Sheet name.
        cell : str
            Cell address.
        formula : str
            Formula string including the leading '=' (e.g. '=SUM(B5:B10)').
        """
        ws = self._get_sheet(sheet)
        self._validate_cell_address(cell)
        self._check_not_protected_sheet(sheet)

        if not formula.startswith("="):
            raise ExcelToolError(
                f"Formula must start with '='. Got: {formula!r}"
            )

        current = ws[cell].value
        ws[cell] = formula
        self._log("write_formula", sheet, cell, current, formula)

    def batch_write(self, operations: list[dict]) -> list[dict]:
        """
        Write multiple cells in a single call.

        Parameters
        ----------
        operations : list of dicts, each with keys:
            - sheet (str)
            - cell (str)
            - value (Any)
            - formula (str, optional — use instead of value for formula writes)
            - force (bool, optional — default False)

        Returns
        -------
        List of result dicts with keys: sheet, cell, status, error (if any).

        Example
        -------
        tool.batch_write([
            {"sheet": "Model", "cell": "H5", "value": "AcmeCorp"},
            {"sheet": "Model", "cell": "I109", "value": 85.0},
            {"sheet": "Model", "cell": "J6",  "formula": "=I6*1.1"},
        ])
        """
        results = []
        for op in operations:
            sheet = op.get("sheet")
            cell = op.get("cell")
            try:
                if "formula" in op:
                    self.write_formula(sheet, cell, op["formula"])
                else:
                    self.write_cell(sheet, cell, op["value"], force=op.get("force", False))
                results.append({"sheet": sheet, "cell": cell, "status": "ok"})
            except ExcelToolError as e:
                results.append({"sheet": sheet, "cell": cell, "status": "error", "error": str(e)})
        return results

    def format_cell(
        self,
        sheet: str,
        cell: str,
        number_format: Optional[str] = None,
        bold: Optional[bool] = None,
        italic: Optional[bool] = None,
        font_color: Optional[str] = None,
        bg_color: Optional[str] = None,
        horizontal_align: Optional[str] = None,
        border_style: Optional[str] = None,
        font_size: Optional[float] = None,
        font_name: Optional[str] = None,
    ) -> None:
        """
        Apply formatting to a cell without changing its value.

        Parameters
        ----------
        sheet : str
            Sheet name.
        cell : str
            Cell address.
        number_format : str, optional
            Excel number format string (e.g. '#,##0.0', '0.0%', '$#,##0').
        bold : bool, optional
        italic : bool, optional
        font_color : str, optional
            Hex color string without '#' (e.g. 'FF0000' for red).
        bg_color : str, optional
            Hex fill color string without '#' (e.g. 'FFFF00' for yellow).
        horizontal_align : str, optional
            One of: 'left', 'center', 'right', 'general'.
        border_style : str, optional
            One of: 'thin', 'medium', 'thick', 'dashed', 'dotted'.
        font_size : float, optional
        font_name : str, optional
            Font name (e.g. 'Calibri', 'Arial').
        """
        ws = self._get_sheet(sheet)
        self._validate_cell_address(cell)
        self._check_not_protected_sheet(sheet)

        c = ws[cell]

        # Font
        font_kwargs = {}
        if bold is not None:
            font_kwargs["bold"] = bold
        if italic is not None:
            font_kwargs["italic"] = italic
        if font_color is not None:
            font_kwargs["color"] = font_color
        if font_size is not None:
            font_kwargs["size"] = font_size
        if font_name is not None:
            font_kwargs["name"] = font_name
        if font_kwargs:
            existing = c.font
            c.font = Font(
                name=font_name or existing.name,
                size=font_size or existing.size,
                bold=bold if bold is not None else existing.bold,
                italic=italic if italic is not None else existing.italic,
                color=font_color or (existing.color.rgb if existing.color else None),
            )

        # Fill
        if bg_color is not None:
            c.fill = PatternFill(fill_type="solid", fgColor=bg_color)

        # Alignment
        if horizontal_align is not None:
            c.alignment = Alignment(horizontal=horizontal_align)

        # Number format
        if number_format is not None:
            c.number_format = number_format

        # Border
        if border_style is not None:
            side = Side(style=border_style)
            c.border = Border(left=side, right=side, top=side, bottom=side)

        self._log("format", sheet, cell, None, {
            "number_format": number_format,
            "bold": bold,
            "font_color": font_color,
            "bg_color": bg_color,
        })

    def save(self) -> Path:
        """
        Save the working copy to output_path.

        Returns
        -------
        Path to the saved file.
        """
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self._wb.save(self.output_path)
        self._log("save", None, None, None, str(self.output_path))
        return self.output_path

    def get_audit_log(self, as_json: bool = False) -> Union[list[dict], str]:
        """
        Return the full audit log.

        Parameters
        ----------
        as_json : bool
            If True, return a JSON string. Otherwise return a list of dicts.
        """
        if as_json:
            return json.dumps(self._audit, indent=2, default=str)
        return self._audit

    def save_audit_log(self, path: Union[str, Path, None] = None) -> Path:
        """
        Write the audit log to a JSON file.

        Parameters
        ----------
        path : str | Path, optional
            Destination path. Defaults to output_path with '_audit.json' suffix.
        """
        if path is None:
            path = self.output_path.with_name(self.output_path.stem + "_audit.json")
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(self._audit, indent=2, default=str), encoding="utf-8")
        return path

    def list_sheets(self) -> list[str]:
        """Return all sheet names in the workbook."""
        return self._wb.sheetnames

    def get_sheet_info(self, sheet: str) -> dict:
        """Return basic metadata about a sheet."""
        ws = self._get_sheet(sheet)
        return {
            "name": sheet,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
        }

    def is_formula_cell(self, sheet: str, cell: str) -> bool:
        """Return True if the cell currently contains a formula."""
        ws = self._get_sheet(sheet)
        self._validate_cell_address(cell)
        value = ws[cell].value
        return isinstance(value, str) and value.startswith("=")

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _get_sheet(self, sheet: str):
        if sheet not in self._wb.sheetnames:
            valid = ", ".join(f'"{s}"' for s in self._wb.sheetnames)
            raise ExcelToolError(
                f'Sheet "{sheet}" not found. Valid sheets: {valid}'
            )
        return self._wb[sheet]

    def _validate_cell_address(self, cell: str) -> None:
        try:
            col_str, row = coordinate_from_string(cell)
            column_index_from_string(col_str)  # validates column letters
            if row < 1:
                raise ValueError
        except Exception:
            raise ExcelToolError(
                f'Invalid cell address: "{cell}". Expected format like "B5" or "AA100".'
            )

    def _check_not_protected_sheet(self, sheet: str) -> None:
        if sheet in _PROTECTED_SHEETS:
            raise ExcelToolError(
                f'Sheet "{sheet}" is protected and cannot be modified.'
            )

    def _log(
        self,
        operation: str,
        sheet: Optional[str],
        cell: Optional[str],
        old_value: Any,
        new_value: Any,
    ) -> None:
        self._audit.append({
            "timestamp": datetime.utcnow().isoformat() + "Z",
            "agent_id": self.agent_id,
            "operation": operation,
            "sheet": sheet,
            "cell": cell,
            "old_value": old_value,
            "new_value": new_value,
        })
